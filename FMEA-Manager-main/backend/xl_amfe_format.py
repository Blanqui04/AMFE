import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def apply_standard_formatting(workbook: openpyxl.Workbook):

    for sheet in workbook.worksheets:
        if sheet.title != "Combined Data":
            continue

        max_col = 31  # AE

        # --- Column widths (manual sizes) ---
        column_widths = {
            "A": 3, "B": 9.29, "C": 27.86, "D": 27.86, "E": 30.71, "F": 34.14, "G": 30.71,
            "H": 38, "I": 34.14, "J": 24.43, "K": 32.14, "L": 3, "M": 34, "N": 3, "O": 26,
            "P": 3, "Q": 3, "R": 5, "S": 32.14, "T": 16.43, "U": 13.57, "V": 13.57,
            "W": 9.29, "X": 19.29, "Y": 13.57, "Z": 3, "AA": 3, "AB": 3, "AC": 5,
            "AD": 3, "AE": 13.57
        }
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width

        # --- Row heights ---
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 22.5
        sheet.row_dimensions[3].height = 75

        # --- Specific cell content ---
        sheet["B3"].value = "Historial"

        # --- Row 1: Main title ---
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        c = sheet.cell(row=1, column=1)
        c.value = "Anàlisi modal de fallades i efectes del procés (AMFE de procés)"
        c.font = Font(name="Arial", bold=True, size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        c.border = Border(*[Side(style='thin', color='000000')]*4)

        # --- Row 2: Stage headings ---
        stage_defs = [
            ((1, 2), "MILLORA CONTINUA", "B7DEE8"),
            ((3, 5), "2a Etapa: Anàlisi estructural", "C6E0B4"),
            ((6, 8), "3a Etapa: Anàlisi funcional", "F8CBAD"),
            ((9, 11), "4a Etapa: Anàlisi de fallades", "FFD966"),
            ((12, 18), "5a Etapa: Anàlisi de riscos", "D9E1F2"),
            ((19, 31), "6a Etapa: Optimització", "D9E1F2"),
        ]

        for idx, ((start, end), title, color) in enumerate(stage_defs):
            if start > max_col:
                continue
            end = min(end, max_col)
            sheet.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            c = sheet.cell(row=2, column=start)
            if idx == 0:
                c.value = title
                c.font = Font(name="Arial", bold=True, size=8)  # Bold, size 8
            else:
                c.value = title
                c.font = Font(name="Arial", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            c.border = Border(*[Side(style='thin', color='000000')]*4)

        for col in range(1, max_col+1):
            c = sheet.cell(row=2, column=col)
            if c.value is None:
                c.font = Font(name="Arial", bold=True, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = Border(*[Side(style='thin', color='000000')]*4)

        # --- Row 3: Headers ---
        def fmt(r, col, fill, rotate=False, align_bottom=False):
            if col > max_col:
                return
            cell = sheet.cell(row=r, column=col)
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            cell.font = Font(name="Arial", bold=False, size=9)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom" if align_bottom else "center",
                textRotation=90 if rotate else 0,
                wrap_text=True
            )
            cell.border = Border(*[Side(style='thin', color='000000')]*4)

        for col in [1, 2]:
            fmt(3, col, "FFF2CC")
        for col in [3, 6, 9]:
            fmt(3, col, "DCDCDC")
        for col in [4, 7, 10]:
            fmt(3, col, "D7E7F7")
        for col in [5, 8, 11]:
            fmt(3, col, "D7AFFF")
        for col in [12, 14, 16, 17, 18, 26, 27, 28, 29, 30]:
            fmt(3, col, "A9D08E", rotate=True, align_bottom=True)
        for col in list(range(12, 23)) + list(range(29, 32)):
            if col not in [12, 14, 16, 17, 18, 26, 27, 28, 29, 30]:
                fmt(3, col, "A9D08E")
        for col in range(21, 26):
            fmt(3, col, "F2F2F2")

        # --- Global formatting ---
        sheet.freeze_panes = sheet["A4"]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=max_col):
            for cell in row:
                # Row 1
                if cell.row == 1:
                    cell.font = Font(name="Arial", bold=True, size=13)
                # Row 2
                elif cell.row == 2:
                    if cell.column in [1, 2]:
                        cell.font = Font(name="Arial", size=8)
                    else:
                        cell.font = Font(name="Arial", bold=True, size=10)
                # Row 3
                elif cell.row == 3:
                    cell.font = Font(name="Arial", bold=True, size=9)
                # All other rows
                else:
                    cell.font = Font(name="Arial", size=9)
                # Date columns V (22) and Y (25): short date format, centered
                if cell.column in [22, 25]:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Alignment
                elif cell.row >= 4:
                    if cell.column in [1, 12, 14, 16, 17, 18, 26, 27, 28, 29, 30]:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    if cell.alignment is None:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal or "center",
                            vertical=cell.alignment.vertical or "center",
                            textRotation=cell.alignment.textRotation if hasattr(cell.alignment, 'textRotation') else 0,
                            wrap_text=True
                        )
                if cell.border is None:
                    cell.border = Border(*[Side(style='thin', color='000000')]*4)

        # --- Borders ---
        thin = Side(style='thin', color='000000')
        medium_thick = Side(style='medium', color='000000')
        thick_cols = [3, 6, 9, 12, 19]  # C, F, I, L, S (after B, E, H, K, R)
        first_data_row = 2
        last_data_row = sheet.max_row
        # Find last row with any data in columns 1 to max_col
        for r in range(sheet.max_row, first_data_row-1, -1):
            if any(sheet.cell(row=r, column=col).value not in (None, "") for col in range(1, max_col+1)):
                last_data_row = r
                break
        # Helper to check if a row is empty
        def is_row_empty(r):
            return all(sheet.cell(row=r, column=col).value in (None, "") for col in range(1, max_col+1))
        # Apply borders
        for r in range(1, last_data_row+1):
            for c in range(1, max_col+1):
                cell = sheet.cell(row=r, column=c)
                # Default: thin borders
                left = thin
                right = thin
                top = thin
                bottom = thin
                # Exterior borders
                if r == 1:
                    top = medium_thick
                if r == last_data_row:
                    bottom = medium_thick
                if c == 1:
                    left = medium_thick
                if c == max_col:
                    right = medium_thick
                # Thick border between rows of header
                if r == 2:
                    top = medium_thick
                if r == 3:
                    top = medium_thick
                if r == 4:
                    top = medium_thick
                # Thick border between data and empty rows
                if r >= first_data_row:
                    if not is_row_empty(r) and r > first_data_row and is_row_empty(r-1):
                        top = medium_thick
                    if is_row_empty(r) and r > first_data_row and not is_row_empty(r-1):
                        top = medium_thick
                # Thick border at the bottom of last data row
                if r == last_data_row:
                    bottom = medium_thick
                # Thick vertical borders between specified columns (from row 4 to last data row)
                if r >= first_data_row and r <= last_data_row and c in thick_cols:
                    left = medium_thick
                # Set border
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
