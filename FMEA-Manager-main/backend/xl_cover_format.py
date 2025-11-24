#import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from pathlib import Path


def format_process_graphics_sheet(workbook, image_name="logo_some.png"):
    project_root = Path(__file__).resolve().parents[1]  # Up from backend/
    image_path = project_root / "assets" / "images" / image_name

    if not image_path.exists():
        raise FileNotFoundError(f"Image not found: {image_path}")
    sheet_name = "Process Graphics"
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        return

    sheet = workbook[sheet_name]

    # --- Column widths ---
    column_widths = {
        "A": 3, "B": 12, "C": 19, "D": 16, "E": 8,
        "F": 8, "G": 16, "H": 16, "I": 12, "J": 3
    }
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # --- Row heights ---
    for i in range(1, 100):
        sheet.row_dimensions[i].height = 14.25
    sheet.row_dimensions[14].height = 4.5

    # --- Merge and format A1:D4 ---
    sheet.merge_cells("A1:D4")
    cell = sheet["A1"]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(name="Arial", size=20, bold=True)

    # --- Insert image at G1 ---
    img = Image(image_path)
    img.width = 283  # Approx. 7.85 * 36
    img.height = 65  # Approx. 1.8 * 36
    sheet.add_image(img, "G1")

    # --- Label cells ---
    label_style = {
        "font": Font(name="Arial", size=10, bold=True),
        "alignment": Alignment(horizontal="left", vertical="center")
    }

    labels = {
        "C7": "Customer Part Nº",
        "C9": "Drawing Nº",
        "C11": "Engineering lev.",
        "C13": "Last rev. date",
        "G7": "SOME Part Nº"
    }

    for cell_address, text in labels.items():
        cell = sheet[cell_address]
        cell.value = text
        cell.font = label_style["font"]
        cell.alignment = label_style["alignment"]

    # --- G9:H9 -> Part description (centered, bold) ---
    sheet.merge_cells("G9:H9")
    g9 = sheet["G9"]
    g9.value = "Part description"
    g9.font = Font(name="Arial", size=10, bold=True)
    g9.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("G10:H11")
    sheet["G10"].alignment = Alignment(horizontal="center", vertical="center")  # or as needed

    # --- Statement title in B17:I17 ---
    sheet.merge_cells("B17:I17")
    b17 = sheet["B17"]
    b17.value = "SOME hereby confirms the process FMEA is conform to the following statements."
    b17.font = Font(name="Arial", size=10, bold=True)
    b17.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Statement body in B19:I25 ---
    sheet.merge_cells("B19:I25")
    statement_text = (
        "· For high (H) risk, based on AP, measures are introduced and marked with the current status of completion.\n\n"
        "· For possible failures, in case of the defined measures are not introduced and/or \n"
        "  effectiveness is not proved, there are special measures identified.\n\n"
        "· For possible failures not identified nor evaluated measures will be taken upon findings to minimize risks"
    )
    b19 = sheet["B19"]
    b19.value = statement_text
    b19.font = Font(name="Arial", size=10, bold=False)
    b19.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # --- Thin black line borders between row 4-5 and 15-16 (A to J) ---
    thin_border = Border(bottom=Side(style="thin", color="000000"))
    for col in range(1, 11):  # A to J
        sheet.cell(row=4, column=col).border = thin_border
        sheet.cell(row=15, column=col).border = thin_border

    # --- Footer section ---
    # B48:D48 - 'Processes analyzed:'
    sheet.merge_cells("B48:D48")
    b48 = sheet["B48"]
    b48.value = "Processes analyzed:"
    b48.font = Font(name="Arial", size=10, bold=True)
    b48.alignment = Alignment(horizontal="left", vertical="center")

    # E48:I48 - 'Comments:'
    sheet.merge_cells("E48:I48")
    e48 = sheet["E48"]
    e48.value = "Comments:"
    e48.font = Font(name="Arial", size=10, bold=True)
    e48.alignment = Alignment(horizontal="left", vertical="center")

    # E49:J58 - blank comment box
    sheet.merge_cells("E49:J58")
    e49 = sheet["E49"]
    e49.value = ""
    e49.alignment = Alignment(wrap_text=True)

    # B63 - 'Date:'
    sheet["B63"].value = "Date:"
    sheet["B63"].font = Font(name="Arial", size=10, bold=True)
    sheet["B63"].alignment = Alignment(horizontal="left")

    # B65 - 'Name:'
    sheet["B65"].value = "Name:"
    sheet["B65"].font = Font(name="Arial", size=10, bold=True)
    sheet["B65"].alignment = Alignment(horizontal="left")

    # B67 - 'Position:'
    sheet["B67"].value = "Position:"
    sheet["B67"].font = Font(name="Arial", size=10, bold=True)
    sheet["B67"].alignment = Alignment(horizontal="left")

    # E61:I62 - 'signature with company stamp'
    sheet.merge_cells("E61:I62")
    sig_cell = sheet["E61"]
    sig_cell.value = "Signature with company stamp"
    sig_cell.font = Font(name="Arial", size=10, bold=True)
    sig_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # E63:I68 - blank for signature
    sheet.merge_cells("E63:I68")
    sheet["E63"].value = ""

    # --- Page setup ---
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_margins.left = 0.7
    sheet.page_margins.right = 0.7
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75

    # --- Print area up to column J and row 68 ---
    sheet.print_area = "A1:J68"
