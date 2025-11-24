"""
AMFE Manager: Core logic for AMFE file processing and analysis
"""

import os
import re
import zipfile
import logging
import openpyxl
import warnings
import openpyxl.styles
from datetime import datetime
from matplotlib.figure import Figure
from openpyxl.chart.label import DataLabelList
from typing import List, Dict, Optional, Tuple
from openpyxl.chart import PieChart, Reference
from utilities.logger import setup_logging

from .errors import AMFEFileError, AMFEProcessError
from .xl_amfe_format import apply_standard_formatting
from .xl_cover_format import format_process_graphics_sheet

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

logger = setup_logging("amfe")
# Ensure logs folder exists
log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)

# Create log file name with timestamp
log_filename = os.path.join(
    log_dir, f"amfe_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
)

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,  # Or INFO in production
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_filename, mode="w", encoding="utf-8"),
        logging.StreamHandler(),  # Also prints to console
    ],
)

logger = logging.getLogger(__name__)

NOM_CARPETA_AMFES = (
    r"\\Server\some\1002-SPS\3.5 PROCESS ASSURANCE_P-FMEA\4 Draft\00_AMFES en curs"
)
CARPETA_AMFE = r"\03_AMFE"


class AmfeManager:
    def __init__(self, base_folder: str = NOM_CARPETA_AMFES):
        try:
            logger.debug(f"Initializing AMFE Manager with base folder: {base_folder}")
            self.base_folder = base_folder
            self.available_processes = self.scan_available_processes()
            logger.debug(f"Found processes: {self.available_processes}")
        except Exception as e:
            logger.exception("Error initializing AMFE Manager")
            raise AMFEFileError(f"Error initializing AMFE Manager: {e}")

    def scan_available_processes(self) -> List[str]:
        try:
            logger.debug(f"Scanning available processes in: {self.base_folder}")
            return [
                d
                for d in os.listdir(self.base_folder)
                if os.path.isdir(os.path.join(self.base_folder, d))
            ]
        except Exception as e:
            logger.exception("Error scanning processes")
            raise AMFEProcessError(f"Error scanning processes: {e}")

    def get_amfe_paths(self, selected_processes: List[str]) -> List[str]:
        logger.debug(f"Getting AMFE paths for selected processes: {selected_processes}")
        amfe_paths = []
        for proc in selected_processes:
            proc_folder = os.path.join(self.base_folder, proc)
            amfe_folder = os.path.join(proc_folder, CARPETA_AMFE.strip("\\/"))
            if os.path.exists(amfe_folder):
                logger.debug(f"Found AMFE folder: {amfe_folder}")
                amfe_paths.append(amfe_folder)
            else:
                logger.warning(f"AMFE folder not found for process: {proc}")
        return amfe_paths

    def get_amfe_excels(self, paths: List[str]) -> List[str]:
        logger.debug(f"Looking for AMFE Excel files in paths: {paths}")
        amfe_excels = []
        for path in paths:
            try:
                files = [
                    f
                    for f in os.listdir(path)
                    if f.endswith(".xlsx") and not f.startswith("~$")
                ]
                for file in files:
                    full_path = os.path.join(path, file)
                    logger.debug(f"Found AMFE Excel: {full_path}")
                    amfe_excels.append(full_path)
            except Exception as e:
                logger.warning(f"Error listing files in {path}: {e}")
        return amfe_excels

    def is_row_empty(self, row) -> bool:
        return all(cell is None or cell == "" for cell in row)

    def is_valid_excel(self, file_path: str) -> bool:
        if not os.path.exists(file_path):
            logger.warning(f"File does not exist: {file_path}")
            return False
        if not zipfile.is_zipfile(file_path):
            logger.warning(f"Not a valid Excel zip file: {file_path}")
            return False
        try:
            with zipfile.ZipFile(file_path) as zf:
                return "xl/workbook.xml" in zf.namelist()
        except Exception as e:
            logger.warning(f"Error validating Excel file {file_path}: {e}")
            return False

    def read_amfe_data(self, file_path: str, include_header: bool) -> List[list]:
        if not self.is_valid_excel(file_path):
            logger.warning(f"Invalid Excel skipped: {file_path}")
            return []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            logger.debug(f"Reading workbook: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to open workbook: {e}")
            return []
        try:
            sheet_name = "P-FMEA_AIAG-VDA"
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                return []
            sheet = wb[sheet_name]
            data_rows = []
            max_row = sheet.max_row
            columns_to_keep = list(range(1, 19)) + list(range(22, 35))
            if include_header:
                for row_idx in [1, 7, 8]:
                    row = [
                        sheet.cell(row=row_idx, column=c).value for c in range(1, 35)
                    ]
                    row_trimmed = [row[i - 1] for i in columns_to_keep]
                    data_rows.append(row_trimmed)
            current_row = 10
            while current_row <= max_row:
                row_data = [
                    sheet.cell(row=current_row, column=c).value for c in range(1, 35)
                ]
                trimmed_row = [row_data[i - 1] for i in columns_to_keep]
                if self.is_row_empty(trimmed_row):
                    break
                data_rows.append(trimmed_row)
                current_row += 1
                logger.debug(f"Extracted {len(data_rows)} rows from {file_path}")
            return data_rows
        except Exception as e:
            logger.warning(f"Error reading sheet: {e}")
            return []
        finally:
            wb.close()

    def write_process_names(
        self, sheet, process_names: List[str], start_row: int = 50, start_col: int = 2
    ):
        for i, name in enumerate(process_names):
            cell = sheet.cell(row=start_row + i, column=start_col)
            cell.value = name

    def clean_process_names(self, names: List[str]) -> List[str]:
        cleaned = []
        for name in names:
            name = re.sub(
                r"^\d{2}_", "", name
            )  # Step 1: Remove leading index (e.g., '01_', '10_')
            name = name.replace("_", " ")  # Step 2: Replace underscores with spaces
            name = name.replace("+", " and ")  # Step 3: Replace '+' with 'and'
            name = re.sub(r"\s+", " ", name).strip()  # Step 4: Normalize spacing
            name = re.sub(
                r"\s+(I{1,3}|IV|V|VI{0,3}|IX|X)$", "", name
            )  # Step 5: Remove trailing Roman numerals like ' I', ' II' (outside parentheses only)

            # Step 6: Capitalize words, skipping words in parentheses
            def smart_cap(word):
                if word.lower() == "and":
                    return "and"
                if word.startswith("(") and word.endswith(")"):
                    return word
                return word.capitalize()

            name_parts = name.split(" ")
            name = " ".join(smart_cap(word) for word in name_parts)

            cleaned.append(name)
        return cleaned

    def count_risk_levels(
        self, all_rows: List[list]
    ) -> Tuple[Dict[str, int], Dict[str, int]]:
        before_counts = {"H": 0, "M": 0, "L": 0}
        after_counts = {"H": 0, "M": 0, "L": 0}
        col_before = 16
        col_after = 29
        for row in all_rows[3:]:
            if row is None or all(cell is None for cell in row):
                continue
            if len(row) > col_before:
                value = row[col_before]
                risk_before = str(value).strip().upper() if value else ""
                if risk_before in before_counts:
                    before_counts[risk_before] += 1
            if len(row) > col_after:
                value = row[col_after]
                risk_after = str(value).strip().upper() if value else ""
                if risk_after in after_counts:
                    after_counts[risk_after] += 1
        return before_counts, after_counts

    def get_risk_summary_data(
        self, filepath: str, show: str = "before"
    ) -> Dict[str, int]:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = (
                wb["Combined Data"] if "Combined Data" in wb.sheetnames else wb.active
            )
            all_rows = [row for row in sheet.iter_rows(values_only=True)]
            before_counts, after_counts = self.count_risk_levels(all_rows)
            return before_counts if show == "before" else after_counts
        except Exception:
            return {"H": 0, "M": 0, "L": 0}

    def create_pie_chart(self, sheet, title, data, table_position, chart_position):
        # Desired order: M, H, L
        risk_order = [
            ("Medium Risk (M)", "M"),
            ("High Risk (H)", "H"),
            ("Low Risk (L)", "L"),
        ]
        table_row, table_col = table_position
        chart_row, chart_col = chart_position

        # Write headers for the table
        sheet.cell(row=table_row, column=table_col, value="Risk Level")
        sheet.cell(row=table_row, column=table_col + 1, value="Count")

        # Write table data
        for i, (label, key) in enumerate(risk_order):
            sheet.cell(row=table_row + 1 + i, column=table_col, value=label)
            sheet.cell(
                row=table_row + 1 + i, column=table_col + 1, value=data.get(key, 0)
            )

        # Create pie chart
        chart = PieChart()
        chart.style = 10
        chart.width = 9.6
        chart.height = 8

        labels_ref = Reference(
            sheet, min_col=table_col, min_row=table_row + 1, max_row=table_row + 3
        )
        data_ref = Reference(
            sheet, min_col=table_col + 1, min_row=table_row + 1, max_row=table_row + 3
        )

        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        chart.title = title
        chart.legend.position = "b"

        # Configure labels: show only percent
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLeaderLines = False

        # Add chart to sheet at specified location
        anchor_cell = f"{openpyxl.utils.get_column_letter(chart_col)}{chart_row}"
        sheet.add_chart(chart, anchor_cell)
        return chart

    def display_risk_pie_chart(
        self, data: Dict[str, int], title: str = "Risk Distribution"
    ) -> Figure:
        labels = ["High Risk (H)", "Medium Risk (M)", "Low Risk (L)"]
        values = [data.get("H", 0), data.get("M", 0), data.get("L", 0)]
        colors = ["#C40000", "#FFA600", "#00BB00"]
        fig = Figure(figsize=(5, 5), facecolor="none")
        ax = fig.add_subplot(111)
        ax.axis("off")
        total = sum(values)
        if total > 0:
            wedges, texts, autotexts = ax.pie(
                values,
                colors=colors,
                autopct=lambda p: f"{p:.1f}%" if p > 0 else "",
                startangle=90,
                textprops={"fontsize": 10, "color": "white", "fontweight": "bold"},
                wedgeprops={"edgecolor": "black", "linewidth": 0.5},
                pctdistance=0.7,
            )
            ax.set_title(title, fontsize=10, fontweight="bold", pad=5)
            ax.legend(
                wedges,
                labels,
                loc="upper center",
                bbox_to_anchor=(0.5, -0.05),
                frameon=False,
                fontsize=8,
                ncol=3,
            )
            fig.tight_layout(pad=1.0)
        else:
            ax.text(
                0.5,
                0.5,
                "No Risk Data Available",
                ha="center",
                va="center",
                fontsize=12,
            )
        return fig

    def combine_amfe_data(
        self, amfe_paths: List[str], output_file: str
    ) -> Optional[Dict[str, object]]:
        logger.info(f"Combining AMFE data from: {amfe_paths}")
        amfe_files = self.get_amfe_excels(amfe_paths)
        logger.debug(f"Total AMFE files found: {len(amfe_files)}")
        all_rows = []
        valid_files = 0
        for i, file_path in enumerate(amfe_files):
            include_header = valid_files == 0
            file_data = self.read_amfe_data(file_path, include_header)
            if file_data:
                valid_files += 1
                all_rows.extend(file_data)
                if i < len(amfe_files) - 1:
                    num_cols = len(file_data[0])
                    all_rows.append([None] * num_cols)
                else:
                    logger.warning(f"No data read from: {file_path}")

        if valid_files == 0:
            logger.error("No valid AMFE files processed. Aborting combination.")
            return None

        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            graphics_sheet = wb.create_sheet("Process Graphics")
            combined_sheet = wb.create_sheet("Combined Data")

            # Write combined data
            for row_index, row in enumerate(all_rows, 1):
                for col_index, value in enumerate(row, 1):
                    combined_sheet.cell(row=row_index, column=col_index, value=value)

            before_counts, after_counts = self.count_risk_levels(all_rows)

            title_cell = graphics_sheet["A1"]
            title_cell.value = "P-FMEA COVER SHEET"
            title_cell.font = openpyxl.styles.Font(bold=True, size=16)

            self.create_pie_chart(
                graphics_sheet,
                "Initial Risk Distribution (Before Actions)",
                before_counts,
                (2, 12),
                (29, 2),
            )
            self.create_pie_chart(
                graphics_sheet,
                "Residual Risk Distribution (After Actions)",
                after_counts,
                (8, 12),
                (29, 6),
            )

            for col in range(1, 15):
                col_letter = openpyxl.utils.get_column_letter(col)
                graphics_sheet.column_dimensions[col_letter].width = 15

            raw_process_names = [
                os.path.basename(os.path.dirname(p)) for p in amfe_paths
            ]
            logger.debug(f"Raw process names: {raw_process_names}")
            cleaned_names = self.clean_process_names(raw_process_names)
            logger.debug(f"Cleaned process names: {cleaned_names}")
            self.write_process_names(
                graphics_sheet, cleaned_names, start_row=50, start_col=2
            )

            apply_standard_formatting(wb)
            format_process_graphics_sheet(wb)
            wb.save(output_file)
            logger.info(f"Successfully saved combined AMFE file to: {output_file}")

            return {
                "output_file": output_file,
                "before_counts": before_counts,
                "after_counts": after_counts,
            }
        except Exception as e:
            logger.exception(f"Failed to create combined AMFE file: {e}")
            return None
